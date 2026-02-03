import sys
import os

def install_and_import(package):
    import importlib
    try:
        importlib.import_module(package)
    except ImportError:
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", package])
    finally:
        globals()[package] = importlib.import_module(package)

try:
    import pandas as pd
except ImportError:
    print("Pandas not found. Attempting to inspect with openpyxl directly if available, or just failing gracefully.")
    try:
        import openpyxl
    except ImportError:
        print("openpyxl not found either.")

try:
    from docx import Document
except ImportError:
    print("python-docx not found.")
    Document = None

excel_path = r"C:\Code\Harshi-App\build\data\subjects\physics.xlsx"
docx_path = r"D:\ChatGpt\Harshi\Comprehensive study guide\1.4.1 Comprehensive Guide Format for Grade 8 Physics.docx"

print(f"--- Inspecting Excel: {excel_path} ---")
if os.path.exists(excel_path):
    try:
        # Try reading with pandas
        try:
            df = pd.read_excel(excel_path)
            print("Columns:", df.columns.tolist())
            print("First 2 rows:")
            print(df.head(2).to_string())
        except Exception as e:
            print(f"Pandas read failed: {e}")
            # Fallback to openpyxl
            import openpyxl
            wb = openpyxl.load_workbook(excel_path)
            sheet = wb.active
            print("OpenPyXL - Sheet Names:", wb.sheetnames)
            print("OpenPyXL - First Row (Values):")
            for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
                print(row)
    except Exception as e:
        print(f"Failed to read Excel: {e}")
else:
    print("Excel file not found.")

print(f"\n--- Inspecting Docx: {docx_path} ---")
if os.path.exists(docx_path):
    if Document:
        try:
            doc = Document(docx_path)
            full_text = []
            for para in doc.paragraphs:
                full_text.append(para.text)
            print("Docx Content Sample (first 1000 chars):")
            print("\n".join(full_text)[:1000])
        except Exception as e:
            print(f"Failed to read Docx: {e}")
    else:
        print("python-docx library not available to read docx.")
else:
    print("Docx file not found.")
