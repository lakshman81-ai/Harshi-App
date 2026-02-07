#!/usr/bin/env python3
"""
Excel to CSV Conversion Script for Harshi-App Restructuring
Converts Excel files to CSV format organized by Subject/Topic
"""

import openpyxl
import csv
import os
from pathlib import Path
from typing import Dict, List

# Base paths
BASE_DIR = Path(__file__).parent.parent
DATA_DIR = BASE_DIR / 'public' / 'data'
QUESTIONNAIRE_DIR = BASE_DIR / 'public' / 'questionnaire'
STUDYGUIDE_DIR = BASE_DIR / 'public' / 'studyguide'
HANDOUT_DIR = BASE_DIR / 'public' / 'Handout'

def clean_cell_value(cell):
    """Clean cell value, handle None and empty strings"""
    if cell is None:
        return ""
    value = str(cell).strip()
    return value if value != "None" else ""

def read_excel_sheet(file_path, sheet_name):
    """Read Excel sheet and return as list of dictionaries"""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            print(f"  [WARN] Sheet '{sheet_name}' not found in {file_path.name}")
            return []
        
        ws = wb[sheet_name]
        data = []
        
        # Get header row
        headers = [clean_cell_value(cell.value) for cell in ws[1]]
        
        # Read data rows
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers) and headers[i]:
                    row_dict[headers[i]] = clean_cell_value(value)
            if any(row_dict.values()):  # Skip empty rows
                data.append(row_dict)
        
        wb.close()
        return data
    except Exception as e:
        print(f"  [ERROR] Error reading {file_path.name}/{sheet_name}: {e}")
        return []

def write_csv(file_path, data, fieldnames):
    """Write data to CSV file"""
    if not data:
        print(f"  [WARN] No data to write to {file_path}")
        return
    
    # Create parent directory
    file_path.parent.mkdir(parents=True, exist_ok=True)
    
    try:
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
            writer.writeheader()
            writer.writerows(data)
        print(f"  [+] Created: {file_path.relative_to(BASE_DIR)}")
    except Exception as e:
        print(f"  [ERROR] Error writing {file_path}: {e}")

def get_subject_topic_mapping():
    """Read master file and get subject/topic mapping"""
    master_file = DATA_DIR / 'StudyHub_Master.xlsx'
    
    # Read Subjects
    subjects_data = read_excel_sheet(master_file, 'Subjects')
    subjects = {s['subject_key']: s['name'] for s in subjects_data if 'subject_key' in s and 'name' in s}
    
    # Read Topics
    topics_data = read_excel_sheet(master_file, 'Topics')
    
    # Build mapping: {subject_key: {topic_id: topic_name}}
    mapping = {}
    for topic in topics_data:
        if 'subject_key' not in topic or 'topic_id' not in topic or 'topic_name' not in topic:
            continue
        subject_key = topic['subject_key']
        topic_id = topic['topic_id']
        topic_name = topic['topic_name']
        
        if subject_key not in mapping:
            mapping[subject_key] = {}
        mapping[subject_key][topic_id] = topic_name
    
    return subjects, mapping

def create_master_indices(subjects, mapping):
    """Create _master_index.csv for each content type"""
    print("\n[*] Creating master index files...")
    
    index_data = []
    for subject_key, topics in mapping.items():
        subject_name = subjects.get(subject_key, subject_key)
        for topic_id, topic_name in topics.items():
            index_data.append({
                'subject_key': subject_key,
                'subject_name': subject_name,
                'topic_id': topic_id,
                'topic_name': topic_name,
                'topic_folder': topic_name.replace(' ', '_').replace("'", "")
            })
    
    fieldnames = ['subject_key', 'subject_name', 'topic_id', 'topic_name', 'topic_folder']
    
    # Write master indices
    write_csv(QUESTIONNAIRE_DIR / '_master_index.csv', index_data, fieldnames)
    write_csv(STUDYGUIDE_DIR / '_master_index.csv', index_data, fieldnames)
    write_csv(HANDOUT_DIR / '_master_index.csv', index_data, fieldnames)

def convert_quiz_questions():
    """Convert Quiz_Questions from all subject files"""
    print("\n[*] Converting Quiz Questions...")
    
    subjects_dir = DATA_DIR / 'subjects'
    
    for subject_file in subjects_dir.glob('*.xlsx'):
        subject_key = subject_file.stem  # e.g., 'physics'
        print(f"\n  Processing {subject_key}...")
        
        quiz_data = read_excel_sheet(subject_file, 'Quiz_Questions')
        
        # Group by topic_id
        by_topic = {}
        for question in quiz_data:
            topic_id = question.get('topic_id', '')
            if not topic_id:
                continue
            if topic_id not in by_topic:
                by_topic[topic_id] = []
            by_topic[topic_id].append(question)
        
        # Write CSV per topic
        for topic_id, questions in by_topic.items():
            # Get topic name from questions or use topic_id
            topic_name = questions[0].get('topic_name', topic_id) if questions else topic_id
            topic_folder = topic_name.replace(' ', '_').replace("'", "")
            
            output_path = QUESTIONNAIRE_DIR / subject_key.title() / topic_folder / 'questions.csv'
            
            fieldnames = ['question_id', 'topic_id', 'question_text', 'option_a', 'option_b', 
                         'option_c', 'option_d', 'correct_answer', 'explanation', 'difficulty', 'hint', 'xp_reward']
            
            write_csv(output_path, questions, fieldnames)

def convert_study_content():
    """Convert Study_Content, Formulas, Key_Terms from all subject files"""
    print("\n[*] Converting Study Content...")
    
    subjects_dir = DATA_DIR / 'subjects'
    
    for subject_file in subjects_dir.glob('*.xlsx'):
        subject_key = subject_file.stem
        print(f"\n  Processing {subject_key}...")
        
        # Read all study-related sheets
        study_content = read_excel_sheet(subject_file, 'Study_Content')
        formulas = read_excel_sheet(subject_file, 'Formulas')
        key_terms = read_excel_sheet(subject_file, 'Key_Terms')
        
        # Group by topic_id
        by_topic = {}
        
        # Add study content
        for item in study_content:
            topic_id = item.get('topic_id', '')
            if not topic_id:
                continue
            if topic_id not in by_topic:
                by_topic[topic_id] = []
            by_topic[topic_id].append({
                'content_id': item.get('content_id', ''),
                'content_type': item.get('content_type', 'text'),
                'title': item.get('content_title', ''),
                'content': item.get('content_text', ''),
                'url': item.get('video_url', '') or item.get('image_url', ''),
                'order_index': item.get('order_index', '')
            })
        
        # Add formulas as content
        for formula in formulas:
            topic_id = formula.get('topic_id', '')
            if not topic_id:
                continue
            if topic_id not in by_topic:
                by_topic[topic_id] = []
            by_topic[topic_id].append({
                'content_id': formula.get('formula_id', ''),
                'content_type': 'formula',
                'title': formula.get('formula_label', ''),
                'content': formula.get('formula_text', ''),
                'url': '',
                'order_index': formula.get('order_index', '')
            })
        
        # Add key terms
        for term in key_terms:
            topic_id = term.get('topic_id', '')
            if not topic_id:
                continue
            if topic_id not in by_topic:
                by_topic[topic_id] = []
            by_topic[topic_id].append({
                'content_id': term.get('term_id', ''),
                'content_type': 'key_term',
                'title': term.get('term', ''),
                'content': term.get('definition', ''),
                'url': '',
                'order_index': ''
            })
        
        # Write CSV per topic
        for topic_id, content_items in by_topic.items():
            # Get topic name
            topic_name = topic_id.replace('-', ' ').replace('_', ' ').title()
            if content_items and 'topic_name' in content_items[0]:
                topic_name = content_items[0].get('topic_name', topic_name)
            topic_folder = topic_name.replace(' ', '_').replace("'", "")
            
            output_path = STUDYGUIDE_DIR / subject_key.title() / topic_folder / 'content.csv'
            
            fieldnames = ['content_id', 'content_type', 'title', 'content', 'url', 'order_index']
            
            write_csv(output_path, content_items, fieldnames)

def main():
    """Main conversion function"""
    print("[*] Starting Excel to CSV Conversion...")
    print(f"[*] Base Directory: {BASE_DIR}")
    
    # Step 1: Get subject/topic mapping
    print("\n[*] Reading master data...")
    subjects, mapping = get_subject_topic_mapping()
    print(f"  Found {len(subjects)} subjects, {sum(len(t) for t in mapping.values())} topics")
    
    # Step 2: Create master indices
    create_master_indices(subjects, mapping)
    
    # Step 3: Convert quiz questions
    convert_quiz_questions()
    
    # Step 4: Convert study content
    convert_study_content()
    
    print("\n[+] Conversion complete!")
    print(f"\n[*] Output directories:")
    print(f"  - {QUESTIONNAIRE_DIR}")
    print(f"  - {STUDYGUIDE_DIR}")
    print(f"  - {HANDOUT_DIR}")

if __name__ == '__main__':
    main()
